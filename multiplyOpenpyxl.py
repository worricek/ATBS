#! /usr/bin/python3

import openpyxl, sys
from openpyxl.styles import Font, NamedStyle
from openpyxl.utils import get_column_letter

num=int(input('Number to multiply signore e signorina.... '))

wb=openpyxl.Workbook()
sheet=wb.get_active_sheet()

#fontObj=Font(bold=True)
#styleObj=NamedStyle(name='styleObj')
#styleObj.font=fontObj

row=1
for a in range(1,num+1):
    counter=1
    column=1
    while True:
        if row==1 and column==1:
            column=2
            for i in range(1, num+1):
                columnLetter=get_column_letter(column)
                sheet[columnLetter+str(row)]=i
                sheet[columnLetter+str(row)].font=Font(bold=True)
#                sheet[columnLetter+str(row)].style=styleObj
                column+=1
                #print(column, i)    #print the first row
            column=2
            row=2
            for j in range(1, num+1):
                sheet['A'+str(row)]=j
                sheet['A'+str(row)].font=Font(bold=True)
#                sheet['A'+str(row)].style=styleObj
                row+=1
                #print(row, j)    #print the first column
            row=2
            continue
        if column==1:
            column=2
            continue
        while True:
            #print(counter, num)
            if counter>num:
                row+=1
                break
            total=(row-1)*counter
            columnLetter=get_column_letter(column)
            sheet[columnLetter+str(row)]=total                
            #print(counter, total)
            counter+=1
            column+=1
        break
        
#sheet[row+column].value=row+column
wb.save('multiply ' + str(num) + '.xlsx')
