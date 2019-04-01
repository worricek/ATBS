#! /usr/bin/python3

#Opens a file that user inputs. Reads the contents and copies it to a new file (arg 2),
#but inverts the locations. Eg C7 in the old becomes G3 in the new. 

import openpyxl, os, sys
from openpyxl.utils import get_column_letter, column_index_from_string

#TODO Take the input of the file you want to treat and the 2 integers.

file=sys.argv[1]
newSS=sys.argv[2]

#TODO open the stated file and the empty spreadsheet. 

print('Current working directory is ....' + os.getcwd())
wb=openpyxl.load_workbook(file)
newFile=openpyxl.Workbook()

#TODO Read in the contents of the spreadsheet to a list.

newSheet=newFile.get_active_sheet()
sheet=wb.get_active_sheet()
endofSheet=get_column_letter(sheet.max_column) + str(sheet.max_row)

#TODO Write N lines from the list to new spreadsheet.

for rowOfCellObjects in list(sheet['A1':endofSheet]):
    for cellObj in rowOfCellObjects:
        column=get_column_letter(cellObj.row)
        row=str(column_index_from_string(cellObj.column))
        Coords=column+row
#        print(cellObj.row, row, cellObj.column, column, Coords, cellObj.value)
        newSheet[Coords].value=cellObj.value

#TODO Save and close the files.

wb.save('example.xlsx')
newFile.save('example1.xlsx')
