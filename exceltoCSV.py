import csv, openpyxl,  os
from openpyxl.utils import get_column_letter

for excelFile in os.listdir('/home/worrall/excelSpreadsheets'):
    # Skip non-xlsx files, load the workbook object.
    if not excelFile.endswith('.xlsx'):
        continue
    wb=openpyxl.load_workbook('/home/worrall/excelSpreadsheets/'+excelFile)
    for sheetName in wb.get_sheet_names():
        # Loop through every sheet in the workbook.
        sheet = wb.get_sheet_by_name(sheetName)

        # Create the CSV filename from the Excel filename and sheet title.
        outputFile=open('/home/worrall/excelSpreadsheets/csv/'+excelFile+'_'+sheetName+'.csv', 'w')
        
        # Create the csv.writer object for this CSV file.
        outputWriter=csv.writer(outputFile)
        
        # Loop through every row in the sheet.        
        for rowNum in range(1, sheet.max_row + 1):
            rowData = []            # append each cell to this list
            # Loop through each cell in the row.
            for colNum in range(1, sheet.max_column + 1):
                # Append each cell's data to rowData.
                #print(colNum)
                #cell=get_column_letter(colNum)+str(rowNum)
                #print(cell)
                rowData.append(sheet[get_column_letter(colNum)+str(rowNum)].value)
            # Write the rowData list to the CSV file.
            outputWriter.writerow(rowData)
        outputFile.close()
