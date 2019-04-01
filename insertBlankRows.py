#! /usr/bin/python3

#Opens a file that user inputs. Reads the contents and copies it to a new file.
#with empty rows inserted. Takes 2 numbers as input. 
#First (N) is the row where you want to insert the blank lines and the
#Second (M) is the number of blank rows to insert

import openpyxl, os, sys
from openpyxl.utils import get_column_letter, column_index_from_string

#TODO Take the input of the file you want to treat and the 2 integers.

#file=input('Whats the XLSX file to want to mangle? ....')
#insertPlace=int(input('What line number to do you want to insert blank lines? ....'))
#blankLines=int(input('How many blank lines do you want to insert? ......'))

file=sys.argv[1]
insertPlace=int(sys.argv[2])
blankLines=int(sys.argv[3])

#TODO open the stated file and the empty spreadsheet. 

print('Current working directory is ....' + os.getcwd())
wb=openpyxl.load_workbook(file)
newFile=openpyxl.Workbook()

#TODO Read in the contents of the spreadsheet to a list.

newSheet=newFile.get_active_sheet()
sheet=wb.get_active_sheet()
endofSheet=get_column_letter(sheet.max_column) + str(sheet.max_row)

#TODO Write N lines from the list to new spreadsheet.
rowNum=1
for rowOfCellObjects in list(sheet['A1':endofSheet]):
    for cellObj in rowOfCellObjects:
        if rowNum>=insertPlace:
            print('Go time')
            newRow=cellObj.row + blankLines
            print(newRow)
            newCoords=cellObj.column+str(newRow)
            print(newCoords)
            newSheet[newCoords].value=cellObj.value
        else:
            print('Not yet')
            newSheet[cellObj.coordinate].value=cellObj.value
        print(cellObj.coordinate,  cellObj.row, cellObj.column,  cellObj.value) # sheet[cellObj.coordinate].value)
    print('--- END OF ROW ---')
    rowNum+=1

#TODO Save and close the files.

wb.save('example.xlsx')
newFile.save('example1.xlsx')
