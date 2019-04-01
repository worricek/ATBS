#! /usr/bin/python3

import openpyxl, os, sys
from openpyxl.utils import get_column_letter, column_index_from_string

#TODO Open blank Excel file

wd=os.chdir('/home/worrall/Downloads/cdr-UDC') #set working directory
newFile=openpyxl.Workbook()
newSheet=newFile.get_active_sheet()

#Loop through each file in the directory

#TODO Loop through file each the directory and open txt files

column=1
for filename in os.listdir(wd):
    row=1
    if '.txt' in os.path.basename(filename):
        file=open(filename, 'r') #Open the file

#TODO for each line in a txt file write this to a new row in the same column

        for line in file.readlines(): #open the file again so we can read each line separately. 
#            print(line)
            coords=get_column_letter(column)+str(row)
#            print(coords)
            newSheet[coords].value=line
#            print(newSheet[coords].value)
            row+=1
        file.close()
        column+=1

#TODO Save the Excel. Job done.

newFile.save('text2Excel.xlsx')
